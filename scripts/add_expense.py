#!/usr/bin/env python3
"""
領収書データを経費Excelに追記する。
引数: 日付 店舗名 品目 金額 勘定科目 [画像パス]
"""
import sys
import os
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PIL import Image

OUTPUT_DIR = Path(__file__).parent.parent / 'output'
EXCEL_PATH = OUTPUT_DIR / '経費管理.xlsx'
IMG_DIR    = OUTPUT_DIR / '領収書画像'

HEADERS = ['No', '日付', '店舗・支払先', '品目・内容', '金額（税込）', '勘定科目', '備考', '画像ファイル名']

def get_or_create_workbook():
    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '経費一覧'
        # ヘッダー
        for col, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(fill_type='solid', fgColor='1F4E79')
            cell.alignment = Alignment(horizontal='center')
        # 列幅
        widths = [6, 12, 20, 25, 15, 15, 20, 25]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    return wb, ws

def compress_image(src_path: str, filename: str) -> str:
    IMG_DIR.mkdir(exist_ok=True)
    dst = IMG_DIR / filename
    img = Image.open(src_path)
    img.thumbnail((800, 800))
    img.save(str(dst), 'JPEG', quality=60)
    return filename

def add_row(date, store, item, amount, account, memo='', img_path=''):
    OUTPUT_DIR.mkdir(exist_ok=True)
    wb, ws = get_or_create_workbook()

    # 連番
    last_row = ws.max_row
    no = last_row if last_row > 1 else 1

    # 画像圧縮
    img_filename = ''
    if img_path and Path(img_path).exists():
        ext_date = date.replace('/', '')
        img_filename = compress_image(img_path, f'receipt_{ext_date}_{no}.jpg')

    row = [no, date, store, item, int(amount), account, memo, img_filename]
    ws.append(row)

    # 交互の背景色
    fill_color = 'DCE6F1' if no % 2 == 0 else 'FFFFFF'
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.fill = PatternFill(fill_type='solid', fgColor=fill_color)
        cell.alignment = Alignment(horizontal='center' if col in [1,2,5,6] else 'left')

    wb.save(EXCEL_PATH)
    print(f'追記完了: {EXCEL_PATH}')
    print(f'  No.{no} | {date} | {store} | {item} | ¥{amount} | {account}')
    if img_filename:
        print(f'  画像: {img_filename}')

if __name__ == '__main__':
    if len(sys.argv) < 6:
        print('Usage: add_expense.py 日付 店舗名 品目 金額 勘定科目 [備考] [画像パス]')
        sys.exit(1)
    date    = sys.argv[1]
    store   = sys.argv[2]
    item    = sys.argv[3]
    amount  = sys.argv[4]
    account = sys.argv[5]
    memo    = sys.argv[6] if len(sys.argv) > 6 else ''
    img     = sys.argv[7] if len(sys.argv) > 7 else ''
    add_row(date, store, item, amount, account, memo, img)
